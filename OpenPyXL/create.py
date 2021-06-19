from openpyxl.workbook import Workbook
# from openpyxl import load_workbook
# Excel: workbook > worksheet

# Create a workbook object/instance
wb = Workbook()

# create an active worksheet
ws = wb.active

# create worksheet title
ws.title = 'Names and Colors'

# create Python list of names
names = ['Dan', 'April', 'Neal']
colors = ['Red', 'Blue', 'Yellow']

ws['A1'] = 'Names'
ws['B1'] = 'Colors'

# add data to ws
starting_row = 2
for index, name in enumerate(names):
	ws.cell(row=starting_row, column=1).value = name
	ws.cell(row=starting_row, column=2).value = colors[index]
	starting_row += 1

# use a formula
ws['B5'] = '=B2&","&B3&","&B4'

# Save our spreadsheet
wb.save('names_colors.xlsx')