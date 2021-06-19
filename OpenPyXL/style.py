from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side

# load existing spreadsheet (saved in the same working directory)
wb = load_workbook('hello.xlsx')
ws = wb.active

cell = ws['A1']

# change the font/color on 'cell'
'''
args:
color="hexcolor code", use google for color pick and remove the "#"
'''
cell.font = Font(
	size=30,
	bold=True,
	italic=False,
	color="82d660"
	)

# Define a Side for our border
'''args
style options:  Value must be one of {'mediumDashDot', 'med
iumDashed', 'slantDashDot', 'hair', 'dashDot', 'mediumD
ashDotDot', 'thin', 'dashed', 'dashDotDot', 'medium', '
double', 'thick', 'dotted'}
'''

my_bd = Side(style='thick', color='d80d0d')
cell = ws['B3']
# apply to the cell
cell.border = Border(
	left=my_bd,
	right=my_bd,
	top=my_bd,
	bottom=my_bd
	)

# save
wb.save('names2.xlsx')
print('File saved.')