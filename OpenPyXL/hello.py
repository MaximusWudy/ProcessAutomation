from openpyxl.workbook import Workbook
from openpyxl import load_workbook
# Excel: workbook > worksheet

# Create a workbook object/instance
# wb = Workbook()

# create an active worksheet
# ws = wb.active

# load existing spreadsheet (saved in the same working directory)
wb = load_workbook('hello.xlsx')
ws = wb.active

# Print something from our spreadsheet
print(ws['A2'].value)
print(f"{ws['A2'].value}: {ws['B2'].value}") # a 'f' string: f"{a} other_str {b}"

# grab a column or row (loop through the tuple to print values)
col_a = ws['A'] # returns a tuple with objects
print(list(map(lambda x: x.value, col_a)))
# OR ues a for loop
print([cell.value for cell in col_a])
# OR better rendering
for cell in col_a:
	print(f"{cell.value}\n")

row_a = ws['1']
print([cell.value for cell in row_a])

# grab ranges of cells
range_a = ws['A2':'A10']
print([cell.value for parent_cell in range_a for cell in parent_cell]) # 2 level tuple

# Iterate through rows
for row in ws.iter_rows(min_row=2, max_row=6, min_col=1, max_col=2, values_only=True):
	print(f"{row[0]}: {row[1]}")
'''
Args:
min_row: start with row number
max_row: stop with row number
values_only: only return value instead of instances
'''

# Iterate through cols
for col in ws.iter_cols(min_row=1, max_row=6, min_col=2, max_col=2, values_only=True):
	print(col)

# Change and save data
ws["A2"] = "Johnny"

# create python list of names
names = ['Dan', 'April', 'Neal']
colors = ['Red', 'Blue', 'Yellow']
# change many cells
starting_row = 12
for index, name in enumerate(names):
	ws.cell(row=starting_row, column=1).value = name
	ws.cell(row=starting_row, column=2).value = colors[index]
	starting_row += 1

wb.save('new_hello.xlsx')
print("File saved.")
